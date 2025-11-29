"""
Step 2 – Fill Excel invoice from SpringAhead JSON and export to PDF (Windows),
or fill an .xlsx invoice file (non-Windows).

Requirements:
    Windows COM path:
        pip install pywin32

    Cross-platform path (.xlsx only):
        pip install openpyxl

Files expected in the same folder as this script:
    - springahead_current_week.json   (output of Step 1)
    - INVOICE (Template).xlsx         (your invoice template)
"""

import os
import json
import calendar
import re
import sys
from datetime import datetime, timedelta
import subprocess
import shutil

# -------- Platform detection --------
IS_WINDOWS = sys.platform.startswith("win")

# Try to import pywin32 only on Windows
win32 = None
if IS_WINDOWS:
    try:
        import win32com.client as win32  # type: ignore
        import win32timezone  # noqa: F401
    except ImportError:
        win32 = None  # pywin32 not available; we'll fall back to openpyxl

# openpyxl is used for the portable backend
try:
    from openpyxl import load_workbook
except ImportError as e:
    # On Windows we can still run via COM; on non-Windows we *require* openpyxl
    if not IS_WINDOWS:
        raise RuntimeError(
            "openpyxl is required to run the invoice step on this platform.\n\n"
            "Install it with:\n    pip install openpyxl"
        ) from e
    # On Windows, COM path might still work, so we don't hard-fail here
    load_workbook = None  # type: ignore


def get_app_root():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


SCRIPT_DIR = get_app_root()
JSON_PATH = os.path.join(SCRIPT_DIR, "springahead_current_week.json")
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "INVOICE (Template).xlsx")


# ---------- Helpers ----------


def load_entries_from_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    entries = data.get("entries", [])

    def parse_date(e):
        return datetime.strptime(e["date"], "%m/%d/%Y")

    return sorted(entries, key=parse_date)


def detect_period_string(entries):
    """
    Build period string like '11 - 1 al 15 - 2025'
    based on month/year of the entries and which half of the month.
    """
    if not entries:
        raise ValueError("No entries in JSON; cannot compute period.")

    dt0 = datetime.strptime(entries[0]["date"], "%m/%d/%Y")
    month = dt0.month
    year = dt0.year

    max_day = max(datetime.strptime(e["date"], "%m/%d/%Y").day for e in entries)
    if max_day <= 15:
        start_day, end_day = 1, 15
    else:
        start_day, end_day = 16, calendar.monthrange(year, month)[1]

    period_str = f"{month} - {start_day} al {end_day} - {year}"
    return period_str


def safe_filename(name: str) -> str:
    """Remove characters that are invalid in Windows filenames."""
    return re.sub(r'[\\/*?:"<>|]', "_", name)


def parse_consultant_name(full_name):
    """
    Parse hispanic full names

    Rules:
      - First word = first name
      - If second word is an initial (e.g. 'O.') → treat it as middle name
      - First last name = next available word after middle names
    """
    full_name = full_name.strip()
    parts = full_name.split()
    if len(parts) < 1:
        raise ValueError("Name cannot be empty.")

    first_name = parts[0]

    # Default assumption
    first_last = None

    # Detect middle-name initial (e.g. O.)
    if len(parts) >= 3 and re.fullmatch(r"[A-Za-z]\.", parts[1]):
        # Skip the initial → last name starts at parts[2]
        first_last = parts[2]
    else:
        # Otherwise, second word IS the first last name
        if len(parts) >= 2:
            first_last = parts[1]
        else:
            first_last = first_name  # fallback only if no last name exists

    initial = first_name[0].upper()
    short_name = f"{initial}. {first_last}"
    return full_name, short_name


def compute_time_blocks(total_hours):
    """
    Given total hours from SpringAhead (e.g. 8.00, 8.25, 7.5),
    return:
      - morning_from, morning_to
      - afternoon_from, afternoon_to

    Base:
      morning:  7:00 AM -> 11:00 AM (4h)
      afternoon: 12:00 PM -> 4:00 PM (4h)
    Overtime/undertime adjusts only the afternoon 'To' time in 15-minute increments.
    """
    base_hours = 8.0
    dummy_date = datetime(2000, 1, 1)
    morning_from = dummy_date.replace(hour=7, minute=0)
    morning_to = dummy_date.replace(hour=11, minute=0)
    afternoon_from = dummy_date.replace(hour=12, minute=0)
    afternoon_base_to = dummy_date.replace(hour=16, minute=0)

    diff_hours = float(total_hours) - base_hours
    diff_minutes = diff_hours * 60.0

    quarter = 15
    diff_minutes_rounded = int(round(diff_minutes / quarter) * quarter)

    afternoon_to = afternoon_base_to + timedelta(minutes=diff_minutes_rounded)

    def fmt(dt):
        return dt.strftime("%I:%M %p").lstrip("0")  # e.g. "7:00 AM"

    return (
        fmt(morning_from),
        fmt(morning_to),
        fmt(afternoon_from),
        fmt(afternoon_to),
    )


def resolve_consultant_name(get_cell_value, set_cell_value):
    """
    Shared logic for resolving the consultant name.

    get_cell_value / set_cell_value are backend-specific callables
    operating on Excel cell B6.
    """
    existing_name = get_cell_value() or ""
    existing_name = str(existing_name).strip()

    # First priority: env var (set by GUI)
    full_name_env = os.getenv("SPRINGAHEAD_FULL_NAME", "").strip()

    if full_name_env:
        full_name_input = full_name_env
        set_cell_value(full_name_input)  # write into template for future runs
    elif existing_name:
        full_name_input = existing_name
    else:
        # No env var and no name in template.
        # If we have a real TTY (CLI), we can prompt.
        if sys.stdin is not None and sys.stdin.isatty():
            full_name_input = input(
                "Enter your full name (first + one or two last names): "
            ).strip()
            while not full_name_input:
                full_name_input = input(
                    "Name cannot be empty. Please enter full name: "
                ).strip()
            set_cell_value(full_name_input)
        else:
            raise RuntimeError(
                "Consultant name is missing.\n\n"
                "Please do ONE of the following and run the app again:\n"
                "- Fill the 'Consultant full name' field in the GUI, or\n"
                "- Type your name into cell B6 of the invoice template."
            )

    return parse_consultant_name(full_name_input)


# ---------- Backend: Windows COM Excel + PDF ----------


def run_step2_windows(entries, period_str):
    if win32 is None:
        raise RuntimeError(
            "pywin32 (win32com.client) is not available on this system.\n"
            "Install it with:\n    pip install pywin32\n\n"
            "Or run this script on a platform where openpyxl is available."
        )

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = True  # set False if you want headless

    wb = excel.Workbooks.Open(TEMPLATE_PATH)
    ws = wb.Worksheets(1)  # assume first sheet is the invoice

    consultant_cell = ws.Cells(6, 2)  # B6

    def get_cell_value():
        return consultant_cell.Value

    def set_cell_value(val):
        consultant_cell.Value = val

    full_name, short_name = resolve_consultant_name(get_cell_value, set_cell_value)

    try:
        # ----- Invoice Number (merged E4:F4 → anchor E4) -----
        invoice_cell = ws.Cells(4, 5)  # E4
        current_number = invoice_cell.Value
        if current_number is None:
            current_number = 0
        try:
            current_number = int(current_number)
        except Exception:
            current_number = 0
        new_number = current_number + 1
        invoice_cell.Value = new_number

        # ----- Period (merged E5:F5 → anchor E5) -----
        period_cell = ws.Cells(5, 5)  # E5
        period_cell.Value = period_str

        # ----- Clear only A–D rows 9–38 -----
        first_data_row = 9
        last_data_row = 38
        for r in range(first_data_row, last_data_row + 1):
            ws.Cells(r, 1).Value = None  # A: Date
            ws.Cells(r, 2).Value = None  # B: From
            ws.Cells(r, 3).Value = None  # C: To
            ws.Cells(r, 4).Value = None  # D: Task

        # ----- Fill rows from JSON entries -----
        current_row = first_data_row

        for entry in entries:
            if current_row + 1 > last_data_row:
                print("Warning: not enough rows in template to fit all entries.")
                break

            date_str = entry["date"]  # e.g. "11/2/2025"
            hours_val = float(entry["hours"])
            dt = datetime.strptime(date_str, "%m/%d/%Y")

            m_from, m_to, a_from, a_to = compute_time_blocks(hours_val)

            # Morning row
            ws.Cells(current_row, 1).Value = dt
            ws.Cells(current_row, 2).Value = m_from
            ws.Cells(current_row, 3).Value = m_to
            ws.Cells(current_row, 4).Value = "Remote IT Support"

            # Afternoon row
            ws.Cells(current_row + 1, 1).Value = dt
            ws.Cells(current_row + 1, 2).Value = a_from
            ws.Cells(current_row + 1, 3).Value = a_to
            ws.Cells(current_row + 1, 4).Value = "Remote IT Support"

            current_row += 2

        # ----- Export to PDF -----
        pdf_filename = safe_filename(f"{short_name} INV ({period_str}).pdf")
        pdf_path = os.path.join(SCRIPT_DIR, pdf_filename)

        wb.Save()

        try:
            ws.ExportAsFixedFormat(
                Type=0,  # PDF
                Filename=pdf_path,
                Quality=0,
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False,
            )
            print(f"Invoice filled and exported to PDF:\n  {pdf_path}")
        except Exception as e:
            print("Export to PDF failed.")
            print(f"Target path: {pdf_path}")
            print(f"Error: {e}")
            print("Leaving Excel open so you can try exporting manually.")
            return

    finally:
        wb.Close(SaveChanges=True)
        excel.Quit()

def try_convert_with_libreoffice(xlsx_path, short_name, period_str):
    """
    Attempt to convert the generated .xlsx invoice to PDF using LibreOffice/soffice.

    This is best-effort:
      - If LibreOffice isn't installed, we just print a message and keep the .xlsx.
      - If conversion fails, we print the error and keep the .xlsx.
    """
    output_dir = os.path.dirname(xlsx_path)

    # Look for a suitable LibreOffice binary
    cmd = None
    for candidate in ("soffice", "libreoffice"):
        if shutil.which(candidate):
            cmd = candidate
            break

    if cmd is None:
        print(
            "LibreOffice was not found on PATH; skipping automatic PDF export "
            "on this platform."
        )
        return

    # We'll ask LibreOffice to write the PDF into the same folder as the .xlsx
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    desired_pdf_name = safe_filename(f"{short_name} INV ({period_str}).pdf")
    desired_pdf_path = os.path.join(output_dir, desired_pdf_name)

    print("\nAttempting automatic PDF export via LibreOffice...")
    try:
        result = subprocess.run(
            [
                cmd,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                output_dir,
                xlsx_path,
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        if result.returncode != 0:
            print("LibreOffice PDF conversion failed. Output:")
            if result.stdout:
                print(result.stdout)
            if result.stderr:
                print(result.stderr)
            print("Leaving the .xlsx invoice as-is.")
            return

        # LibreOffice names the PDF like "<basename>.pdf"
        generated_pdf = os.path.join(output_dir, base_name + ".pdf")
        if os.path.exists(generated_pdf):
            # Rename/move to our desired filename if different
            if os.path.abspath(generated_pdf) != os.path.abspath(desired_pdf_path):
                os.replace(generated_pdf, desired_pdf_path)
            else:
                desired_pdf_path = generated_pdf

            print("Automatic PDF export via LibreOffice succeeded:")
            print(f"  {desired_pdf_path}")
        else:
            print(
                "LibreOffice reported success but the expected PDF was not found.\n"
                "Leaving the .xlsx invoice as-is."
            )

    except Exception as e:
        print("LibreOffice PDF conversion raised an exception; leaving .xlsx only.")
        print(f"Error: {e}")

# ---------- Backend: openpyxl (cross-platform .xlsx) ----------


def run_step2_portable(entries, period_str):
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to generate the invoice on this platform.\n"
            "Install it with:\n    pip install openpyxl"
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.worksheets[0]  # first sheet

    def get_cell_value():
        return ws["B6"].value

    def set_cell_value(val):
        ws["B6"].value = val

    full_name, short_name = resolve_consultant_name(get_cell_value, set_cell_value)

    # Invoice number (E4)
    invoice_cell = ws["E4"]
    current_number = invoice_cell.value
    if current_number is None:
        current_number = 0
    try:
        current_number = int(current_number)
    except Exception:
        current_number = 0
    new_number = current_number + 1
    invoice_cell.value = new_number

    # Period (E5)
    ws["E5"].value = period_str

    # Clear A–D rows 9–38
    first_data_row = 9
    last_data_row = 38
    for r in range(first_data_row, last_data_row + 1):
        for c in range(1, 5):  # A–D
            ws.cell(row=r, column=c, value=None)

    # Fill rows
    current_row = first_data_row
    for entry in entries:
        if current_row + 1 > last_data_row:
            print("Warning: not enough rows in template to fit all entries.")
            break

        date_str = entry["date"]
        hours_val = float(entry["hours"])
        dt = datetime.strptime(date_str, "%m/%d/%Y")

        m_from, m_to, a_from, a_to = compute_time_blocks(hours_val)

        # Morning row
        ws.cell(row=current_row, column=1, value=dt)
        ws.cell(row=current_row, column=2, value=m_from)
        ws.cell(row=current_row, column=3, value=m_to)
        ws.cell(row=current_row, column=4, value="Remote IT Support")

        # Afternoon row
        ws.cell(row=current_row + 1, column=1, value=dt)
        ws.cell(row=current_row + 1, column=2, value=a_from)
        ws.cell(row=current_row + 1, column=3, value=a_to)
        ws.cell(row=current_row + 1, column=4, value="Remote IT Support")

        current_row += 2

    # Save as .xlsx
    xlsx_filename = safe_filename(f"{short_name} INV ({period_str}).xlsx")
    xlsx_path = os.path.join(SCRIPT_DIR, xlsx_filename)
    wb.save(xlsx_path)

    print("Invoice filled and saved as Excel file:")
    print(f"  {xlsx_path}")

    # Try automatic PDF export via LibreOffice, if available
    try_convert_with_libreoffice(xlsx_path, short_name, period_str)

    print(
        "\nIf no PDF file was reported above, you can still open the .xlsx in "
        "Excel/LibreOffice/Numbers and export to PDF manually."
    )



# ---------- Main dispatcher ----------


def main():
    if not os.path.exists(JSON_PATH):
        raise FileNotFoundError(f"JSON not found: {JSON_PATH}")
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    entries = load_entries_from_json(JSON_PATH)
    period_str = detect_period_string(entries)

    if IS_WINDOWS and win32 is not None:
        run_step2_windows(entries, period_str)
    else:
        run_step2_portable(entries, period_str)


if __name__ == "__main__":
    main()
